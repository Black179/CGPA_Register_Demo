import os
import pandas as pd
import json
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import config

class ResultProcessor:
    def __init__(self):
        self.data = None
        self.output_dir = 'output'
        self.timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        self.ensure_output_dir()
    
    def ensure_output_dir(self):
        """Create output directory if it doesn't exist"""
        if not os.path.exists(self.output_dir):
            os.makedirs(self.output_dir)
    
    def load_data(self):
        """Load data based on the configured data source"""
        source_type = config.DATA_SOURCE['type'].lower()
        
        try:
            if source_type == 'csv':
                self.data = pd.read_csv(config.DATA_SOURCE['path'])
            elif source_type == 'excel':
                self.data = pd.read_excel(config.DATA_SOURCE['path'])
            elif source_type == 'database':
                # For database connection, you'll need to install appropriate DB connector
                # Example for MySQL: pip install mysql-connector-python
                import mysql.connector
                conn = mysql.connector.connect(**config.DATA_SOURCE['database'])
                query = "SELECT * FROM student_results"
                self.data = pd.read_sql(query, conn)
                conn.close()
            elif source_type == 'api':
                import requests
                response = requests.get(
                    config.DATA_SOURCE['api']['url'],
                    headers=config.DATA_SOURCE['api'].get('headers', {})
                )
                response.raise_for_status()
                self.data = pd.DataFrame(response.json())
            else:
                raise ValueError(f"Unsupported data source type: {source_type}")
            
            # Ensure required columns exist
            required_columns = [
                'S.No', 'Section', 'Register Number', 'Student Name',
                'SEM 1: Arrear Count', 'SEM 1: Total Arrear', 'SEM 1: Total', 'SEM 1: SGPA',
                'SEM 2: Arrear Count', 'SEM 2: Total Arrear', 'SEM 2: Total', 'SEM 2: SGPA',
                'CGPA (Upto 2nd Semester)', 'Total (Out of 475)', 'Total Arrear Count', 'Signature'
            ]
            
            # Add missing columns with default values
            for col in required_columns:
                if col not in self.data.columns:
                    self.data[col] = ""
            
            # Reorder columns
            self.data = self.data[required_columns]
            
            return True
            
        except Exception as e:
            print(f"Error loading data: {str(e)}")
            return False
    
    def generate_excel(self):
        """Generate Excel report"""
        if self.data is None:
            print("No data available. Please load data first.")
            return None
        
        try:
            output_file = os.path.join(self.output_dir, f"result_analysis_{self.timestamp}.xlsx")
            
            # Create Excel writer
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                self.data.to_excel(
                    writer,
                    sheet_name=config.EXCEL_CONFIG['sheet_name'],
                    index=False,
                    startrow=1
                )
                
                # Get workbook and worksheet objects
                workbook = writer.book
                worksheet = writer.sheets[config.EXCEL_CONFIG['sheet_name']]
                
                # Apply styles
                self._apply_excel_styles(workbook, worksheet)
                
            print(f"Excel report generated: {output_file}")
            return output_file
            
        except Exception as e:
            print(f"Error generating Excel: {str(e)}")
            return None
    
    def _apply_excel_styles(self, workbook, worksheet):
        """Apply formatting to Excel worksheet"""
        # Set column widths
        for i, width in enumerate(config.EXCEL_CONFIG['column_widths'], 1):
            worksheet.column_dimensions[get_column_letter(i)].width = width
        
        # Create styles
        header_fill = PatternFill(
            start_color=config.EXCEL_CONFIG['header_style']['fill']['start_color'],
            fill_type=config.EXCEL_CONFIG['header_style']['fill']['fill_type']
        )
        header_font = Font(
            bold=config.EXCEL_CONFIG['header_style']['font']['bold'],
            color=config.EXCEL_CONFIG['header_style']['font']['color']
        )
        
        # Apply header styles
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal=config.EXCEL_CONFIG['header_style']['alignment']['horizontal'],
                vertical=config.EXCEL_CONFIG['header_style']['alignment']['vertical']
            )
        
        # Apply data styles
        for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
            for cell in row:
                cell.alignment = Alignment(
                    horizontal=config.EXCEL_CONFIG['data_style']['alignment']['horizontal'],
                    vertical=config.EXCEL_CONFIG['data_style']['alignment']['vertical']
                )
    
    def generate_pdf(self):
        """Generate PDF report"""
        if self.data is None:
            print("No data available. Please load data first.")
            return None
        
        try:
            output_file = os.path.join(self.output_dir, f"result_analysis_{self.timestamp}.pdf")
            
            # Create PDF document
            doc = SimpleDocTemplate(
                output_file,
                pagesize=landscape(letter),
                title=config.PDF_CONFIG['title'],
                author=config.PDF_CONFIG['author']
            )
            
            # Prepare data for PDF table
            data = [list(self.data.columns)]  # Header row
            data.extend(self.data.values.tolist())  # Data rows
            
            # Create table
            table = Table(data)
            
            # Add style to table
            style = TableStyle([
                # Header style
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), config.PDF_CONFIG['header_font_size']),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                
                # Data style
                ('FONTNAME', (0, 1), (-1, -1), config.PDF_CONFIG['font_name']),
                ('FONTSIZE', (0, 1), (-1, -1), config.PDF_CONFIG['data_font_size']),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                
                # Grid
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                
                # Row colors
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('E8E8E8')]),
            ])
            
            table.setStyle(style)
            
            # Create elements for PDF
            elements = []
            
            # Add title
            styles = getSampleStyleSheet()
            title_style = ParagraphStyle(
                'Title',
                parent=styles['Title'],
                fontSize=config.PDF_CONFIG['title_font_size'],
                spaceAfter=20,
                alignment=1  # Center
            )
            
            title = Paragraph(config.PDF_CONFIG['title'], title_style)
            elements.append(title)
            
            # Add date
            date_style = ParagraphStyle(
                'Date',
                parent=styles['Normal'],
                fontSize=10,
                spaceAfter=20,
                alignment=2  # Right
            )
            
            date_text = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            date = Paragraph(date_text, date_style)
            elements.append(date)
            
            # Add table
            elements.append(Spacer(1, 10))
            elements.append(table)
            
            # Build PDF
            doc.build(elements)
            
            print(f"PDF report generated: {output_file}")
            return output_file
            
        except Exception as e:
            print(f"Error generating PDF: {str(e)}")
            return None

if __name__ == "__main__":
    # Example usage
    processor = ResultProcessor()
    
    # Load data
    if processor.load_data():
        print(f"Loaded {len(processor.data)} records")
        
        # Generate reports
        excel_file = processor.generate_excel()
        pdf_file = processor.generate_pdf()
        
        if excel_file and pdf_file:
            print("\nReports generated successfully!")
            print(f"- Excel: {os.path.abspath(excel_file)}")
            print(f"- PDF: {os.path.abspath(pdf_file)}")
    else:
        print("Failed to load data. Please check your configuration and try again.")
